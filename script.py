import openpyxl as xl
from PIL import Image
import numpy as np

image_path = 'file.png'

#img = np.array(Image.open(image_path))
# Transparency Issue
png = Image.open(image_path).convert(mode='RGBA')
imgsize = png.size
bg = Image.new('RGB', [max(imgsize)]*2, (255, 255, 255))

bg.paste(png, box=([int((max(imgsize) - i)/2) for i in imgsize]), mask=png.split()[3])
bg = bg.resize(size=(100, 100), resample=Image.LANCZOS)

img = np.array(bg)

# Get 100x100 format
wb = xl.load_workbook('format.xlsx')
ws = wb.active
ws.title = ''.join(image_path.split('.')[:-1])

for y, row in enumerate(img, start=1):
    for x, cell in enumerate(row, start=1):
        # print(xl.utils.cell.get_column_letter(x)+str(y))
        ws[xl.utils.cell.get_column_letter(x)+str(y)].fill = \
            xl.styles.PatternFill(
                fgColor='{:02X}{:02X}{:02X}'.format(*cell),
                fill_type='solid')

wb.save(''.join(image_path.split('.')[:-1]) + '.xlsx')
